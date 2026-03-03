import openpyxl
import json
from collections import Counter

wb = openpyxl.load_workbook('/home/orangels/xm_dev/ls_dev/reportSkillMaker/警情列表_lingao_20241231-20260115_result_case.xlsx', read_only=True)

print("=== Sheet Names ===")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Get headers (first row)
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
    print(f"\nColumn count: {len(headers)}")
    print(f"\nHeaders:")
    for i, h in enumerate(headers):
        print(f"  [{i}] {h}")
    
    # Get first 5 data rows as samples
    print(f"\nSample data (first 5 rows):")
    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        row_count += 1
        print(f"\n  Row {row_count}:")
        for i, val in enumerate(row):
            if val is not None:
                print(f"    [{i}] {headers[i]}: {val}")

wb.close()
