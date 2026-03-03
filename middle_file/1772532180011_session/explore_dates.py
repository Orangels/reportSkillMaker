import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('/home/orangels/xm_dev/ls_dev/reportSkillMaker/警情列表_lingao_20241231-20260115_result_case.xlsx', read_only=True)
ws = wb['警情数据']

headers = None
count = 0
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue
    if count < 5:
        report_time = row[1]
        print(f"Row {i}: type={type(report_time)}, value={repr(report_time)}")
        count += 1

# Check a few rows for date formats
print("\nChecking date parsing...")
from collections import Counter
type_counter = Counter()
sample_strs = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    val = row[1]
    type_counter[type(val).__name__] += 1
    if isinstance(val, str) and len(sample_strs) < 3:
        sample_strs.append(val)

print(f"Type distribution: {dict(type_counter)}")
if sample_strs:
    print(f"String samples: {sample_strs}")

wb.close()
